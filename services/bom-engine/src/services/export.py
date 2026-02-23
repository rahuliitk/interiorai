"""
Multi-format BOM export: Excel (openpyxl), CSV, and PDF (reportlab).

Each exporter returns raw bytes suitable for streaming as a FastAPI response.
"""

from __future__ import annotations

import csv
import io
from datetime import datetime, timezone
from typing import Any

import structlog
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch, mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from src.models.bom import BOMResult

logger = structlog.get_logger(__name__)


# -- Excel export -----------------------------------------------------------

def export_to_excel(bom: BOMResult) -> bytes:
    """Export BOM to an Excel workbook (.xlsx).

    Creates two sheets:
    1. ``BOM Items`` -- all line items with quantities, prices, and totals.
    2. ``Summary`` -- category breakdown and project totals.

    Parameters
    ----------
    bom:
        The complete BOM result to export.

    Returns
    -------
    bytes
        The Excel file contents.
    """
    wb = Workbook()

    # -- Sheet 1: BOM Items -------------------------------------------------
    ws_items = wb.active
    ws_items.title = "BOM Items"

    # Styles
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    currency_format = '#,##0.00'
    percent_format = '0.0%'

    # Title row
    ws_items.merge_cells("A1:I1")
    title_cell = ws_items["A1"]
    title_cell.value = f"Bill of Materials -- {bom.project_id}"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color="2F5496")
    title_cell.alignment = Alignment(horizontal="center")

    ws_items.merge_cells("A2:I2")
    subtitle = ws_items["A2"]
    subtitle.value = (
        f"Room: {bom.room_id} | Variant: {bom.design_variant_id} | "
        f"Generated: {bom.created_at.strftime('%Y-%m-%d %H:%M UTC')}"
    )
    subtitle.font = Font(name="Calibri", italic=True, size=10)
    subtitle.alignment = Alignment(horizontal="center")

    # Headers
    headers = [
        "S.No.", "Category", "Material", "Specification",
        "Qty", "Unit", "Waste %", "Unit Price", "Total (incl. waste)",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_items.cell(row=4, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    for row_idx, item in enumerate(bom.items, start=5):
        serial = row_idx - 4
        estimated = item.estimated_cost or 0.0

        values = [
            serial,
            item.category.value.replace("_", " ").title(),
            item.name,
            item.specification,
            item.quantity,
            item.unit,
            item.waste_factor,
            item.unit_price or 0,
            estimated,
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws_items.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

            if col_idx == 7:
                cell.number_format = percent_format
            elif col_idx in (8, 9):
                cell.number_format = currency_format

    # Total row
    total_row = len(bom.items) + 5
    ws_items.cell(row=total_row, column=8, value="TOTAL").font = Font(bold=True)
    total_cost = bom.summary.total_cost if bom.summary else 0
    total_cell = ws_items.cell(row=total_row, column=9, value=total_cost)
    total_cell.font = Font(bold=True)
    total_cell.number_format = currency_format

    # Column widths
    widths = [6, 16, 30, 35, 10, 8, 10, 14, 18]
    for i, w in enumerate(widths, start=1):
        ws_items.column_dimensions[get_column_letter(i)].width = w

    # -- Sheet 2: Summary ---------------------------------------------------
    ws_summary = wb.create_sheet("Summary")
    ws_summary.merge_cells("A1:D1")
    ws_summary["A1"].value = "BOM Summary"
    ws_summary["A1"].font = Font(name="Calibri", bold=True, size=14, color="2F5496")

    summary_headers = ["Category", "Items", "Subtotal", "% of Total"]
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = ws_summary.cell(row=3, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    if bom.summary:
        for row_idx, cat in enumerate(bom.summary.category_breakdown, start=4):
            ws_summary.cell(
                row=row_idx, column=1,
                value=cat.category.value.replace("_", " ").title(),
            ).border = thin_border
            ws_summary.cell(row=row_idx, column=2, value=cat.item_count).border = thin_border
            sub_cell = ws_summary.cell(row=row_idx, column=3, value=cat.subtotal)
            sub_cell.border = thin_border
            sub_cell.number_format = currency_format
            pct_cell = ws_summary.cell(row=row_idx, column=4, value=cat.percentage_of_total / 100)
            pct_cell.border = thin_border
            pct_cell.number_format = percent_format

    summary_widths = [20, 10, 18, 14]
    for i, w in enumerate(summary_widths, start=1):
        ws_summary.column_dimensions[get_column_letter(i)].width = w

    # Write to bytes
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# -- CSV export -------------------------------------------------------------

def export_to_csv(bom: BOMResult) -> bytes:
    """Export BOM to CSV format.

    Parameters
    ----------
    bom:
        The complete BOM result to export.

    Returns
    -------
    bytes
        The CSV file contents (UTF-8 encoded with BOM for Excel compatibility).
    """
    buf = io.StringIO()
    writer = csv.writer(buf)

    # Header
    writer.writerow([
        "S.No.", "Category", "Material", "Specification",
        "Quantity", "Unit", "Waste Factor", "Unit Price",
        "Total (incl. waste)", "Currency",
    ])

    # Data rows
    for idx, item in enumerate(bom.items, start=1):
        estimated = item.estimated_cost or 0.0
        writer.writerow([
            idx,
            item.category.value.replace("_", " ").title(),
            item.name,
            item.specification,
            item.quantity,
            item.unit,
            f"{item.waste_factor:.1%}",
            f"{item.unit_price or 0:.2f}",
            f"{estimated:.2f}",
            item.currency or "INR",
        ])

    # Totals
    total_cost = bom.summary.total_cost if bom.summary else 0
    writer.writerow([])
    writer.writerow(["", "", "", "", "", "", "", "TOTAL", f"{total_cost:.2f}", "INR"])

    # UTF-8 BOM for Excel compatibility
    content = buf.getvalue()
    return ("\ufeff" + content).encode("utf-8")


# -- PDF export -------------------------------------------------------------

def export_to_pdf(bom: BOMResult) -> bytes:
    """Export BOM to PDF with a professional layout.

    Uses ReportLab to create an A4 landscape document with a title block,
    item table, and summary section.

    Parameters
    ----------
    bom:
        The complete BOM result to export.

    Returns
    -------
    bytes
        The PDF file contents.
    """
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    elements: list[Any] = []

    # Title
    title_style = ParagraphStyle(
        "BOMTitle",
        parent=styles["Title"],
        fontSize=16,
        spaceAfter=6,
        textColor=colors.HexColor("#2F5496"),
    )
    elements.append(Paragraph("Bill of Materials", title_style))

    # Subtitle
    subtitle_style = ParagraphStyle(
        "BOMSubtitle",
        parent=styles["Normal"],
        fontSize=9,
        textColor=colors.grey,
        spaceAfter=12,
    )
    elements.append(Paragraph(
        f"Project: {bom.project_id} | Room: {bom.room_id} | "
        f"Variant: {bom.design_variant_id} | "
        f"Date: {bom.created_at.strftime('%Y-%m-%d %H:%M UTC')}",
        subtitle_style,
    ))
    elements.append(Spacer(1, 6))

    # Items table
    cell_style = ParagraphStyle(
        "CellStyle",
        parent=styles["Normal"],
        fontSize=7,
        leading=9,
    )

    table_data: list[list[Any]] = [
        [
            "S.No.", "Category", "Material", "Specification",
            "Qty", "Unit", "Waste", "Unit Price", "Total",
        ]
    ]

    for idx, item in enumerate(bom.items, start=1):
        estimated = item.estimated_cost or 0.0
        table_data.append([
            str(idx),
            item.category.value.replace("_", " ").title(),
            Paragraph(item.name, cell_style),
            Paragraph(item.specification[:60] if item.specification else "", cell_style),
            f"{item.quantity:.1f}",
            item.unit,
            f"{item.waste_factor:.0%}",
            f"{item.unit_price or 0:,.0f}",
            f"{estimated:,.0f}",
        ])

    # Total row
    total_cost = bom.summary.total_cost if bom.summary else 0
    table_data.append([
        "", "", "", "", "", "", "",
        Paragraph("<b>TOTAL</b>", cell_style),
        Paragraph(f"<b>{total_cost:,.0f}</b>", cell_style),
    ])

    col_widths = [
        30, 65, 120, 150, 40, 35, 35, 55, 60,
    ]

    table = Table(table_data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        # Header
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 8),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        # Data
        ("FONTSIZE", (0, 1), (-1, -1), 7),
        ("ALIGN", (0, 1), (0, -1), "CENTER"),
        ("ALIGN", (4, 1), (4, -1), "RIGHT"),
        ("ALIGN", (6, 1), (8, -1), "RIGHT"),
        # Grid
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        # Alternating rows
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F2F6FC")]),
        # Total row
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#D6E4F0")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        # Padding
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("LEFTPADDING", (0, 0), (-1, -1), 4),
        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 12))

    # Summary section
    if bom.summary and bom.summary.category_breakdown:
        elements.append(Paragraph("Category Summary", title_style))
        elements.append(Spacer(1, 4))

        summary_data: list[list[str]] = [["Category", "Items", "Subtotal", "% of Total"]]
        for cat in bom.summary.category_breakdown:
            summary_data.append([
                cat.category.value.replace("_", " ").title(),
                str(cat.item_count),
                f"{cat.subtotal:,.0f}",
                f"{cat.percentage_of_total:.1f}%",
            ])

        summary_table = Table(summary_data, colWidths=[120, 60, 80, 70])
        summary_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5496")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F6FC")]),
        ]))

        elements.append(summary_table)

    # Footer note
    elements.append(Spacer(1, 12))
    footer_style = ParagraphStyle(
        "Footer",
        parent=styles["Normal"],
        fontSize=7,
        textColor=colors.grey,
    )
    currency = bom.summary.currency if bom.summary else "INR"
    elements.append(Paragraph(
        f"All prices in {currency}. Quantities include waste factor allowances where applicable. "
        f"Generated by OpenLintel BOM Engine.",
        footer_style,
    ))

    doc.build(elements)
    buf.seek(0)
    return buf.read()
